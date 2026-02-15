"""
Ingestão da base de conhecimento: extrai texto do arquivo, divide em chunks, gera embeddings e grava em document_chunks.
Chamado após o upload de documento no platform_backend (ou por script).
Suporta: .txt, .pdf, .xlsx, .xls, .png, .jpg, .jpeg. Requer OPENAI_API_KEY e tabela document_chunks (pgvector).
"""

import base64
import os
from pathlib import Path
from typing import List

# Chunk size e overlap para não estourar contexto e manter continuidade
CHUNK_SIZE = 600
CHUNK_OVERLAP = 80


def _extract_text_excel_xlsx(path: Path) -> str:
    """Extrai texto de planilha .xlsx (todas as abas, todas as células)."""
    from openpyxl import load_workbook
    wb = load_workbook(str(path), read_only=True, data_only=True)
    parts = []
    for sheet in wb.worksheets:
        rows = []
        for row in sheet.iter_rows(values_only=True):
            cells = [str(c).strip() if c is not None else "" for c in row]
            if any(cells):
                rows.append(" | ".join(cells))
        if rows:
            parts.append("\n".join(rows))
    wb.close()
    return "\n\n".join(parts).strip()


def _extract_text_excel_xls(path: Path) -> str:
    """Extrai texto de planilha .xls (formato antigo)."""
    import xlrd
    wb = xlrd.open_workbook(str(path))
    parts = []
    for sheet in wb.sheets():
        rows = []
        for row_idx in range(sheet.nrows):
            cells = [str(sheet.cell_value(row_idx, col_idx)).strip() for col_idx in range(sheet.ncols)]
            if any(cells):
                rows.append(" | ".join(cells))
        if rows:
            parts.append("\n".join(rows))
    return "\n\n".join(parts).strip()


def _extract_text_image(file_path: str) -> str:
    """Extrai texto de imagem (.png, .jpg, .jpeg) via OpenAI Vision (descrição para RAG)."""
    api_key = os.environ.get("OPENAI_API_KEY", "").strip()
    if not api_key:
        raise ValueError("OPENAI_API_KEY é necessário para processar imagens.")
    path = Path(file_path)
    data = path.read_bytes()
    b64 = base64.standard_b64encode(data).decode("ascii")
    suf = path.suffix.lower()
    mime = "image/png" if suf == ".png" else "image/jpeg"
    from openai import OpenAI
    client = OpenAI(api_key=api_key)
    resp = client.chat.completions.create(
        model=os.environ.get("OPENAI_VISION_MODEL", "gpt-4o-mini"),
        messages=[
            {
                "role": "user",
                "content": [
                    {
                        "type": "text",
                        "text": "Descreva o conteúdo desta imagem ou documento de forma textual e detalhada, incluindo texto visível, números, tabelas e informações relevantes, para uso em busca semântica em uma base de conhecimento. Responda apenas com o texto da descrição, sem introduções.",
                    },
                    {
                        "type": "image_url",
                        "image_url": {"url": f"data:{mime};base64,{b64}"},
                    },
                ],
            }
        ],
        max_tokens=1024,
    )
    text = (resp.choices[0].message.content or "").strip()
    return text


def _extract_text(file_path: str) -> str:
    """Extrai texto de .txt, .pdf, .xlsx, .xls, .png, .jpg ou .jpeg."""
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"Arquivo não encontrado: {file_path}")
    suf = path.suffix.lower()
    if suf == ".txt":
        return path.read_text(encoding="utf-8", errors="replace").strip()
    if suf == ".pdf":
        try:
            from pypdf import PdfReader
            reader = PdfReader(str(path))
            parts = []
            for p in reader.pages:
                t = p.extract_text()
                if t:
                    parts.append(t)
            return "\n\n".join(parts).strip()
        except ImportError:
            raise RuntimeError("Para PDF instale: pip install pypdf") from None
    if suf == ".xlsx":
        try:
            return _extract_text_excel_xlsx(path)
        except ImportError:
            raise RuntimeError("Para Excel instale: pip install openpyxl") from None
    if suf == ".xls":
        try:
            return _extract_text_excel_xls(path)
        except ImportError:
            raise RuntimeError("Para Excel .xls instale: pip install xlrd") from None
    if suf in (".png", ".jpg", ".jpeg"):
        return _extract_text_image(file_path)
    raise ValueError(
        f"Formato não suportado: {suf}. Use .txt, .pdf, .xlsx, .xls, .png, .jpg ou .jpeg."
    )


def _chunk_text(text: str, size: int = CHUNK_SIZE, overlap: int = CHUNK_OVERLAP) -> List[str]:
    """Divide texto em blocos com overlap."""
    if not text or not text.strip():
        return []
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    chunks = []
    start = 0
    while start < len(text):
        end = start + size
        chunk = text[start:end]
        if not chunk.strip():
            start = end - overlap
            continue
        chunks.append(chunk.strip())
        start = end - overlap
    return chunks


def _get_connection():
    import psycopg2
    from psycopg2.extras import RealDictCursor
    url = (
        os.environ.get("PLATFORM_DATABASE_URL", "").strip()
        or os.environ.get("DATABASE_URL", "").strip()
    )
    if not url:
        raise ValueError("DATABASE_URL ou PLATFORM_DATABASE_URL não configurado")
    if "supabase.com" in url and "?" not in url:
        url = url + "?sslmode=require"
    return psycopg2.connect(url, cursor_factory=RealDictCursor)


def ingest_document(file_path: str, tenant_id: str, document_id: str) -> int:
    """
    Processa um documento: extrai texto, chunk, gera embeddings, insere em document_chunks.
    document_id = UUID do registro em documents (tabela).
    Retorna o número de chunks inseridos.
    """
    from .knowledge_rag import _embed

    text = _extract_text(file_path)
    chunks = _chunk_text(text)
    if not chunks:
        return 0

    embeddings = _embed(chunks)

    conn = _get_connection()
    try:
        with conn.cursor() as cur:
            for i, (content, emb) in enumerate(zip(chunks, embeddings)):
                vec_str = "[" + ",".join(str(x) for x in emb) + "]"
                cur.execute(
                    """
                    INSERT INTO document_chunks (tenant_id, document_id, chunk_index, content, embedding)
                    VALUES (%s, %s, %s, %s, %s::vector)
                    """,
                    (tenant_id, document_id, i, content, vec_str),
                )
        conn.commit()
        return len(chunks)
    finally:
        conn.close()


def delete_chunks_for_document(document_id: str) -> None:
    """Remove todos os chunks de um documento (chamar ao deletar o documento)."""
    conn = _get_connection()
    try:
        with conn.cursor() as cur:
            cur.execute("DELETE FROM document_chunks WHERE document_id = %s", (document_id,))
        conn.commit()
    finally:
        conn.close()
